#CodeMod from Trinh Nguyen http://www.dangtrinh.com/2013/10/python-convert-csv-to-excel.html
import csv
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def csv_to_excel(csv_path, excel_path):
	csv_file = open(csv_path, encoding = 'utf-8-sig') #ChineseFixUTF8
	csv.register_dialect('comma', delimiter=",")# CSV use , cut
	reader = csv.reader(csv_file, dialect='comma')
	wb = Workbook()
	ws = wb.worksheets[0]
	for row_index, row in enumerate(reader):
		for column_index, cell in enumerate(row):
			column_letter = get_column_letter((column_index + 1))
			ws.cell('%s%s'%(column_letter, (row_index + 1))).value = cell
	wb.save(filename = excel_path)
	csv_file.close()
WPC_PATH = "./01_WPCOutput/"
csv_list = os.listdir(WPC_PATH)
i=1
for j in csv_list:
	print(str(i)+"."+j)
	i+=1
cp = WPC_PATH+csv_list[eval(input("Which file is CSV?\n"))-1]
while cp.find(".csv") < 0:
	cp = WPC_PATH+csv_list[eval(input("Is not CSV file! Which file is ?\n"))-1]
print("Loading file...." + cp)
ep = WPC_PATH+'myexcel.xlsx'
csv_to_excel(cp, ep)
print("Complete file>" + ep)
